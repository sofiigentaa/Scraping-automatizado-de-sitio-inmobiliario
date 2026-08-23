import asyncio
import os
import re
from datetime import datetime
from urllib.parse import urljoin, parse_qs, urlparse, urlencode

from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd
from playwright.async_api import async_playwright

BASE_URL = "https://www.bienesrosario.com/"
ARCHIVO_EXCEL = "propiedades_copia7.xlsx"

# ==========================================
# FUNCIONES AUXILIARES Y LIMPIEZA
# ==========================================

def limpiar_texto(texto):
    if not texto:
        return ""
    return re.sub(r'\s+', ' ', str(texto)).strip()


def quitar_acentos(texto):
    if not texto:
        return ""
    import unicodedata
    return "".join(
        c for c in unicodedata.normalize('NFD', str(texto))
        if unicodedata.category(c) != 'Mn'
    ).lower()


def sanitizar_email(email_raw):
    if not email_raw:
        return "No publicado"
    limpio = re.sub(r'\s+', '', email_raw).lower()
    match = re.search(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', limpio)
    return match.group(0) if match else "No publicado"


def sanitizar_superficie(superficie_raw):
    if not superficie_raw:
        return "No especificado"
    match = re.search(r'(\d+(?:[\.,]\d+)?)', str(superficie_raw))
    return f"{match.group(1)} m²" if match else "No especificado"


def sanitizar_telefono(telefono_raw):
    nums = re.findall(r'\d+', str(telefono_raw))
    if nums:
        cadena = "".join(nums)
        if len(cadena) >= 8:
            return f"Tel/WA: {cadena}"
    return "No publicado"


def sanitizar_cantidad(texto_raw, etiqueta):
    match = re.search(r'\d+', str(texto_raw))
    if match:
        return f"{match.group(0)} {etiqueta}"
    return "No especificado"


# ==========================================
# EXTRACCIÓN DE FICHA DETALLADA
# ==========================================

def extraer_ficha_estricta(soup_detalle, full_url, fecha_copia, localidad_busqueda):
    titulo = "Sin título"
    h1_tag = soup_detalle.select_one(".main-title h1, h1")
    if h1_tag:
        titulo = limpiar_texto(h1_tag.get_text())
    else:
        title_tag = soup_detalle.find("title")
        if title_tag:
            titulo = limpiar_texto(title_tag.get_text().split("-")[0].split("|")[0])

    contenedor_principal = soup_detalle.select_one("#product-detail, .product-detail, .col-md-8, .box-content")
    if not contenedor_principal:
        contenedor_principal = soup_detalle.body

    if contenedor_principal:
        contenedor_copy = BeautifulSoup(str(contenedor_principal), "html.parser")
        for tag in contenedor_copy.select(".box-data, .agency-info, .breadcrumb, script, style, .social-share"):
            tag.decompose()
        texto_crudo = contenedor_copy.get_text()
    else:
        texto_crudo = soup_detalle.get_text()

    texto_unificado = limpiar_texto(texto_crudo)

    if "Descripción" in texto_unificado:
        partes = texto_unificado.split("Descripción", 1)
        if len(partes) > 1 and len(partes[1].strip()) > 20:
            texto_unificado = partes[1].strip()

    match_corte = re.search(r'(expira\s+en\s+90\s*d[ií]as)', texto_unificado, flags=re.IGNORECASE)
    if match_corte:
        fin_idx = match_corte.end()
        descripcion_unica = texto_unificado[:fin_idx].strip()
    else:
        descripcion_unica = texto_unificado.strip()

    descripcion_unica = re.sub(r'^(?:ampliada\s*)+', '', descripcion_unica, flags=re.IGNORECASE).strip()
    descripcion_unica = limpiar_texto(descripcion_unica)

    tipo_operacion = "No especificado"
    tipo_propiedad = "No especificado"

    prop_tag = soup_detalle.find(lambda tag: tag.name == "span" and "Propiedad:" in tag.text)
    if prop_tag and prop_tag.next_sibling:
        tipo_propiedad = limpiar_texto(str(prop_tag.next_sibling))

    breadcrumb_links = soup_detalle.select(".breadcrumb_soporte .breadcrumb li a")
    breadcrumb_texts = [limpiar_texto(a.get_text()) for a in breadcrumb_links]

    if len(breadcrumb_links) >= 3:
        operacion = breadcrumb_texts[1]
        tipo = breadcrumb_texts[2]
        tipo_operacion = f"{tipo} en {operacion}"
        if tipo_propiedad == "No especificado":
            tipo_propiedad = tipo
    elif len(breadcrumb_links) == 2:
        tipo_operacion = breadcrumb_texts[1]

    vendedor = "Dueño Directo"
    agencia_tag = soup_detalle.select_one(".agency-name span, .agency-logo img")
    if agencia_tag:
        vendedor = agencia_tag.get("alt") or agencia_tag.get_text()
    else:
        box_data = soup_detalle.select_one(".box-data .box-text")
        if box_data:
            lineas_v = [l.strip() for l in box_data.get_text("\n").split("\n") if l.strip()]
            if lineas_v:
                vendedor = lineas_v[0]

    vendedor = re.split(r'\b(?:Ciudad|Provincia|Celular|eMail|Contacto|Tel|WA):', vendedor, flags=re.I)[0]
    vendedor = re.sub(r'^[^\w\s]+', '', vendedor)
    vendedor = limpiar_texto(vendedor)

    str_telefono = "No publicado"
    box_contacto = soup_detalle.select_one(".box-data .box-text")
    if box_contacto:
        match_cel = re.search(r'Celular:\s*([\d\s\-]+)', box_contacto.get_text())
        if match_cel:
            str_telefono = sanitizar_telefono(match_cel.group(1))

    if str_telefono == "No publicado":
        ws_link = soup_detalle.select_one("a.btn-contact-whatsapp[href*='wa.me']")
        if ws_link:
            match_num = re.search(r'wa\.me/(\d+)', ws_link.get("href", ""))
            if match_num:
                str_telefono = sanitizar_telefono(match_num.group(0))

    localidad = localidad_busqueda
    direccion = "No especificada"

    texto_evaluacion = quitar_acentos(f"{' '.join(breadcrumb_texts)} {titulo}")
    for subtitulo in soup_detalle.select(".subtitle"):
        if "ubicaci" in subtitulo.get_text().lower():
            siguiente = subtitulo.find_next_sibling("div")
            if siguiente and "box-text" in siguiente.get("class", []):
                texto_evaluacion += " " + quitar_acentos(siguiente.get_text())

    if "roldan" in texto_evaluacion:
        localidad = "Roldán"
    elif "funes" in texto_evaluacion:
        localidad = "Funes"
    elif "rosario" in texto_evaluacion:
        localidad = "Rosario"

    dir_tag = soup_detalle.find(lambda tag: tag.name == "span" and "Dirección:" in tag.text)
    if dir_tag and dir_tag.next_sibling:
        dir_txt = limpiar_texto(str(dir_tag.next_sibling))
        if dir_txt:
            direccion = dir_txt

    if direccion == "No especificada":
        for subtitulo in soup_detalle.select(".subtitle"):
            if "ubicaci" in subtitulo.get_text().lower():
                siguiente = subtitulo.find_next_sibling("div")
                if siguiente and "box-text" in siguiente.get("class", []):
                    texto_ub = limpiar_texto(siguiente.get_text())
                    dir_limpia = re.sub(r'\b(Rosario|Funes|Roldan|Roldán|Santa Fe)\b', '', texto_ub, flags=re.I)
                    dir_limpia = re.sub(r'[\.,\-\s]+$', '', dir_limpia).strip()
                    dir_limpia = re.sub(r'^[\.,\-\s]+', '', dir_limpia).strip()
                    if dir_limpia:
                        direccion = dir_limpia
                    break

    precio = "No especificado"
    for subtitulo in soup_detalle.select(".subtitle"):
        if "precio" in subtitulo.get_text().lower():
            siguiente = subtitulo.find_next_sibling("div")
            if siguiente:
                precio = limpiar_texto(siguiente.get_text())
                break

    dormitorios = "No especificado"
    dorm_tag = soup_detalle.find(lambda tag: tag.name == "span" and "Dormitorios:" in tag.text)
    if dorm_tag and dorm_tag.next_sibling:
        valor_dorm = limpiar_texto(str(dorm_tag.next_sibling))
        if "monoambiente" in valor_dorm.lower():
            dormitorios = "Monoambiente"
        else:
            dormitorios = sanitizar_cantidad(valor_dorm, "Ambientes")

    banos = "No especificado"
    bano_tag = soup_detalle.find(lambda tag: tag.name == "span" and ("Baños:" in tag.text or "Baño:" in tag.text))
    if bano_tag and bano_tag.next_sibling:
        banos = sanitizar_cantidad(str(bano_tag.next_sibling), "Baños")

    superficie = "No especificado"
    sup_tag = soup_detalle.find(lambda tag: tag.name == "span" and "Superficie:" in tag.text)
    if sup_tag and sup_tag.next_sibling:
        superficie = sanitizar_superficie(str(sup_tag.next_sibling))

    if superficie == "No especificado":
        match_sup = re.search(
            r'(?:lote|terreno|sup|superficie)?\s*(?:de)?\s*(\d+(?:[\.,]\d+)?)\s*(?:m2|m\s*2|m²|mts2|metros\s*cuadrados)',
            descripcion_unica,
            re.IGNORECASE
        )
        if match_sup:
            superficie = f"{match_sup.group(1)} m²"

    str_email = "No publicado"
    emails_encontrados = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', str(soup_detalle))
    for mail in emails_encontrados:
        mail_clean = sanitizar_email(mail)
        if mail_clean != "No publicado" and not any(d in mail_clean for d in ["bienesrosario", "schema.org", "sentry", "w3.org", "bootstrap", "cloudflare"]):
            str_email = mail_clean
            break

    return {
        "Título": titulo,
        "URL": full_url,
        "Tipo de Propiedad": tipo_propiedad,
        "Precio": precio,
        "Localidad": localidad,
        "Dirección": direccion,
        "Descripción": descripcion_unica,
        "Dormitorios / Ambientes": dormitorios,
        "Baños": banos,
        "Superficie": superficie,
        "Contacto / Vendedor": vendedor,
        "Teléfono / WhatsApp": str_telefono,
        "Email": str_email,
        "Fecha de Copia": fecha_copia
    }


async def navegar_con_reintento(page, url, reintentos=3, timeout=30000):
    for intento in range(1, reintentos + 1):
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=timeout)
            return True
        except Exception:
            await asyncio.sleep(2)
    return False


async def obtener_html_seguro(page):
    for _ in range(5):
        try:
            await page.wait_for_load_state("domcontentloaded")
            await page.wait_for_load_state("networkidle", timeout=4000)
            return await page.content()
        except Exception:
            await page.wait_for_timeout(800)
    return await page.content()


async def obtener_primer_itmid(page):
    """Devuelve el itmId de la primera propiedad visible en el listado actual."""
    return await page.evaluate("""
        () => {
            const a = document.querySelector("a[href*='showProduct']");
            if (!a) return null;
            const m = a.href.match(/itmId=(\\d+)/);
            return m ? m[1] : null;
        }
    """)


async def seleccionar_ciudad_con_reintento(page, loc, intentos=5, espera_ms=900):
    """
    Selecciona la ciudad en el combo #com_son_1_1 y VERIFICA que la
    selección realmente haya tomado, leyendo de vuelta el texto de la
    opción seleccionada. El combo se carga vía AJAX recién después de
    elegir la provincia, así que un timeout fijo (como se usaba antes)
    a veces no alcanza y la ciudad queda sin seleccionar -> la búsqueda
    sale sin filtro de ciudad y trae TODA la provincia mezclada.
    Reintenta varias veces antes de rendirse.
    """
    # Primero esperamos a que el combo tenga opciones cargadas (además
    # de la opción vacía por defecto)
    try:
        await page.wait_for_function(
            "() => { const s = document.querySelector('#com_son_1_1'); "
            "return s && s.options.length > 1; }",
            timeout=8000,
        )
    except Exception:
        print(f"  ⚠️  El combo de ciudades no cargó opciones a tiempo para '{loc}'.")

    for intento in range(1, intentos + 1):
        await page.evaluate(
            """
            (loc) => {
                const norm = (str) => str.normalize("NFD").replace(/[\\u0300-\\u036f]/g, "").toLowerCase();
                const select = document.querySelector('#com_son_1_1');
                if (select) {
                    for (let opt of select.options) {
                        if (norm(opt.text).includes(norm(loc))) {
                            select.value = opt.value;
                            select.dispatchEvent(new Event('change', { bubbles: true }));
                            break;
                        }
                    }
                }
            }
            """,
            loc,
        )
        await page.wait_for_timeout(espera_ms)

        seleccion_actual = await page.evaluate(
            """
            () => {
                const select = document.querySelector('#com_son_1_1');
                if (!select || select.selectedIndex < 0) return null;
                return select.options[select.selectedIndex].text;
            }
            """
        )

        if seleccion_actual and quitar_acentos(loc) in quitar_acentos(seleccion_actual):
            return True

        print(f"  ⏳ Reintentando selección de ciudad '{loc}' "
              f"(intento {intento}/{intentos}, seleccionado actualmente: '{seleccion_actual}')...")

    return False


async def obtener_total_propiedades(page):
    """Lee el contador '(N propiedades)' que muestra el sitio, si está visible."""
    try:
        texto = await page.evaluate("() => document.body.innerText")
        m = re.search(r'(\d+)\s+propiedades', texto)
        return int(m.group(1)) if m else None
    except Exception:
        return None


async def avanzar_siguiente_pagina(page, offset_actual, tam_pagina=40):
    """
    Pagina llamando específicamente al link jumpToPage(offset) que
    corresponde al SIGUIENTE offset hacia adelante (offset_actual + tam_pagina
    o el próximo disponible mayor a offset_actual).

    Antes se clickeaba "el primer link jumpToPage que aparezca en el DOM",
    pero cuando hay varios controles de paginación (anterior, números de
    página, siguiente) el primero no siempre es "avanzar" — a veces es
    "volver a la página anterior". Eso hacía que el scraper rebotara para
    siempre entre 2 páginas sin darse cuenta (el itmId sí cambiaba al
    volver atrás, así que el chequeo de "avanzó" daba falso positivo).

    Ahora leemos TODOS los offsets disponibles en los links jumpToPage(N)
    del DOM, y clickeamos específicamente el menor que sea mayor al
    offset actual (es decir, el próximo hacia adelante, nunca hacia atrás).
    """
    primer_id_antes = await obtener_primer_itmid(page)

    hrefs = await page.eval_on_selector_all(
        "a[href*='jumpToPage']",
        "els => els.map(e => e.getAttribute('href'))"
    )

    offsets_disponibles = set()
    for href in hrefs:
        if not href:
            continue
        m = re.search(r"jumpToPage\((\d+)\)", href)
        if m:
            offsets_disponibles.add(int(m.group(1)))

    candidatos_adelante = sorted(o for o in offsets_disponibles if o > offset_actual)
    if not candidatos_adelante:
        return None  # no hay páginas hacia adelante: llegamos al final

    offset_destino = candidatos_adelante[0]

    enlace = page.locator(f"a[href*='jumpToPage({offset_destino})']").first
    if await enlace.count() == 0:
        return None

    try:
        await enlace.click(force=True, timeout=5000)
    except Exception:
        return None

    try:
        await page.wait_for_function(
            """
            (idAnterior) => {
                const a = document.querySelector("a[href*='showProduct']");
                if (!a) return false;
                const m = a.href.match(/itmId=(\\d+)/);
                return m && m[1] !== idAnterior;
            }
            """,
            arg=primer_id_antes,
            timeout=8000,
        )
    except Exception:
        return None

    return offset_destino


# ==========================================
# FLUJO PRINCIPAL DE SCRAPING Y EXCEL
# ==========================================

async def run():
    print("🚀 Iniciando scraper de Bienes Rosario...")
    fecha_copia_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    localidades_objetivo = ["Rosario", "Funes", "Roldán"]
    lista_publicaciones = []
    urls_procesadas = set()

    # --- 1. LECTURA PREVIA DEL ARCHIVO EXCEL ---
    df_previo = pd.DataFrame()

    if os.path.exists(ARCHIVO_EXCEL):
        try:
            print(f"📖 Cargando datos previos desde '{ARCHIVO_EXCEL}'...")
            xls = pd.ExcelFile(ARCHIVO_EXCEL)
            sheet_name = "Propiedades" if "Propiedades" in xls.sheet_names else xls.sheet_names[0]
            df_previo = pd.read_excel(ARCHIVO_EXCEL, sheet_name=sheet_name)
            print(f"   └─ Se encontraron {len(df_previo)} registros anteriores.")
        except Exception as e:
            print(f"⚠️ Error al leer '{ARCHIVO_EXCEL}': {e}. Se creará un archivo nuevo.")

    # --- 2. RASPADO DE LA WEB ---
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context()

        for loc in localidades_objetivo:
            print(f"\n🌐 Buscando propiedades en {loc}...")
            page = await context.new_page()

            exito_conexion = await navegar_con_reintento(page, BASE_URL)
            if not exito_conexion:
                print(f"❌ No se pudo conectar a {BASE_URL} para {loc}. Saltando...")
                await page.close()
                continue

            try:
                await page.get_by_role("link", name="BÚSQUEDA AVANZADA").click(force=True, timeout=10000)
                await page.wait_for_timeout(1000)
            except Exception:
                pass

            await page.evaluate("""
                () => {
                    const norm = (str) => str.normalize("NFD").replace(/[\u0300-\u036f]/g, "").toLowerCase();
                    const seleccionar = (id, texto) => {
                        const select = document.querySelector(id);
                        if (select) {
                            for (let opt of select.options) {
                                if (norm(opt.text).includes(norm(texto))) {
                                    select.value = opt.value;
                                    select.dispatchEvent(new Event('change', { bubbles: true }));
                                    break;
                                }
                            }
                        }
                    };
                    seleccionar('#SearchitmVendedor', 'Dueño Directo');
                    seleccionar('#itmTypeOperation', 'Venta');
                    seleccionar('#com_father_1', 'Santa Fe');
                }
            """)

            await page.wait_for_timeout(600)

            exito_ciudad = await seleccionar_ciudad_con_reintento(page, loc)
            if not exito_ciudad:
                print(f"  ❌ No se pudo confirmar la selección de la ciudad '{loc}' "
                      f"después de varios intentos. Salto esta localidad para NO "
                      f"scrapear datos mezclados con otras ciudades. Revisá "
                      f"debug_pagina1_{loc}.html si esto se repite.")
                await page.close()
                continue

            try:
                async with page.expect_navigation(wait_until="domcontentloaded", timeout=30000):
                    await page.evaluate("""
                        () => {
                            const form = document.querySelector('#frm_buscador_modal');
                            if (form) { form.submit(); }
                        }
                    """)
            except Exception:
                pass

            await page.wait_for_timeout(2000)

            total_reportado = await obtener_total_propiedades(page)
            if total_reportado is not None:
                print(f"  ℹ️  El sitio reporta {total_reportado} propiedades para {loc}.")
                if total_reportado > 300:
                    print(f"  🚨 ADVERTENCIA: {total_reportado} es un número muy alto para "
                          f"una sola localidad — es probable que el filtro de ciudad NO haya "
                          f"tomado y esto sea resultado de toda la provincia. Revisá "
                          f"debug_pagina1_{loc}.html antes de confiar en estos datos.")
            contador_intentos_localidad = 0
            contador_extraidas_localidad = 0
            urls_fallidas_localidad = []
            offset_actual = 0

            num_pagina = 1
            while True:
                for _ in range(3):
                    await page.evaluate("window.scrollBy(0, 1000)")
                    await page.wait_for_timeout(400)

                html_contenido = await obtener_html_seguro(page)
                soup = BeautifulSoup(html_contenido, "html.parser")

                # DEBUG: guardamos un snapshot del HTML de la primera página
                # de cada localidad, por si hace falta diagnosticar más.
                if num_pagina == 1:
                    try:
                        with open(f"debug_pagina1_{loc}.html", "w", encoding="utf-8") as f_debug:
                            f_debug.write(html_contenido)
                        print(f"  🐞 [debug] HTML guardado en debug_pagina1_{loc}.html")
                    except Exception:
                        pass

                cantidad_jumptopage = await page.locator("a[href*='jumpToPage']").count()
                print(f"  🔧 [debug] Enlaces 'jumpToPage' detectados: {cantidad_jumptopage} "
                      f"(offset actual: {offset_actual})")

                # Anclamos SOLO en el patrón real de ficha de propiedad:
                # index.php?action=carro/showProduct&itmId=NNNNN
                # (antes se usaba "id=" en minúscula, que nunca matcheaba
                # "itmId=" con I mayúscula, y terminaba agarrando banners
                # publicitarios del sidebar por coincidencia de texto)
                products = soup.find_all("a", href=re.compile(r"action=carro/showProduct"))

                links_validos = []
                itmids_vistos_pagina = set()
                for a_tag in products:
                    href = a_tag["href"]
                    m_id = re.search(r"itmId=(\d+)", href)
                    if not m_id:
                        continue
                    itm_id = m_id.group(1)
                    if itm_id in itmids_vistos_pagina:
                        continue
                    itmids_vistos_pagina.add(itm_id)
                    full_u = href if href.startswith("http") else urljoin(BASE_URL, href)
                    links_validos.append(full_u)

                print(f"  📄 Página {num_pagina} ({loc}): {len(links_validos)} enlaces a propiedades detectados.")

                if not links_validos:
                    print(f"  ⚠️ No se detectaron más propiedades en la página {num_pagina}.")
                    break

                nuevas_en_esta_pagina = 0
                for full_url in links_validos:
                    if full_url in urls_procesadas:
                        continue
                    
                    urls_procesadas.add(full_url)
                    nuevas_en_esta_pagina += 1
                    contador_intentos_localidad += 1

                    nueva_pestana = await context.new_page()
                    extraccion_exitosa = False
                    try:
                        exito_detalle = await navegar_con_reintento(nueva_pestana, full_url, reintentos=2)
                        if not exito_detalle:
                            urls_fallidas_localidad.append((full_url, "no cargó la página tras reintentos"))
                            continue

                        await nueva_pestana.wait_for_timeout(600)

                        try:
                            boton_ver_mail = nueva_pestana.get_by_text("Ver Mail", exact=False)
                            if await boton_ver_mail.count() > 0:
                                await boton_ver_mail.first.click(force=True, timeout=2000)
                                await nueva_pestana.wait_for_timeout(400)
                        except Exception:
                            pass

                        html_detalle = await obtener_html_seguro(nueva_pestana)
                        soup_detalle = BeautifulSoup(html_detalle, "html.parser")

                        propiedad = extraer_ficha_estricta(soup_detalle, full_url, fecha_copia_actual, loc)
                        lista_publicaciones.append(propiedad)
                        contador_extraidas_localidad += 1
                        extraccion_exitosa = True
                        print(f"    ├─ Capturada #{len(lista_publicaciones)}: {propiedad['Título'][:40]}...")
                    except Exception as e:
                        urls_fallidas_localidad.append((full_url, str(e)[:80]))
                    finally:
                        await nueva_pestana.close()

                # Si no se encontró ninguna propiedad nueva en esta iteración, fuerza el corte
                if nuevas_en_esta_pagina == 0:
                    print(f"  ⚠️ Todas las propiedades de la página {num_pagina} ya habían sido procesadas.")

                hay_siguiente = await avanzar_siguiente_pagina(page, offset_actual)

                if hay_siguiente is None:
                    print(f"  🏁 Fin de páginas alcanzado para {loc}.")
                    break

                offset_actual = hay_siguiente
                num_pagina += 1

            if total_reportado is not None:
                if contador_extraidas_localidad < total_reportado:
                    print(f"  ⚠️  ATENCIÓN: {loc} — el sitio reporta {total_reportado} "
                          f"propiedades, se intentaron {contador_intentos_localidad} y se "
                          f"extrajeron con éxito {contador_extraidas_localidad}.")
                else:
                    print(f"  ✅ {loc}: extraídas con éxito {contador_extraidas_localidad} de {total_reportado} reportadas.")

            if urls_fallidas_localidad:
                print(f"  ⚠️  {len(urls_fallidas_localidad)} ficha(s) NO se pudieron extraer en {loc}:")
                for url_fallida, motivo in urls_fallidas_localidad:
                    print(f"     - {url_fallida} ({motivo})")

            await page.close()

        # --- 2.5 VERIFICACIÓN ACTIVA DE URLs GUARDADAS QUE NO APARECIERON ---
        # No asumimos que "no apareció en la búsqueda de hoy" = "ya no existe".
        # Puede ser que el scraping se haya perdido esa página (aunque ya
        # arreglamos la paginación, mejor no confiar ciegamente) o que dejó
        # de cumplir el filtro pero la publicación siga viva. Por eso
        # visitamos cada URL guardada directamente antes de marcarla
        # "No disponible".
        urls_confirmadas_muertas = set()
        if not df_previo.empty:
            col_url_prev = "URL" if "URL" in df_previo.columns else df_previo.columns[0]
            col_estado_prev = "Estado" if "Estado" in df_previo.columns else None
            col_localidad_prev = "Localidad" if "Localidad" in df_previo.columns else None

            # Si querés que también reintente con las que YA estaban marcadas
            # "No disponible" en corridas anteriores (por si fueron
            # republicadas), poné esto en True. Por defecto no las
            # reintenta, para no perder tiempo visitando links ya
            # confirmados como muertos en corridas previas.
            REVISAR_YA_MARCADAS_NO_DISPONIBLE = False

            urls_ya_capturadas_hoy = {p["URL"] for p in lista_publicaciones}
            candidatas = df_previo[~df_previo[col_url_prev].isin(urls_ya_capturadas_hoy)].copy()

            if col_estado_prev:
                estados_a_saltear = ["Rechazada", "Descartada"]
                if not REVISAR_YA_MARCADAS_NO_DISPONIBLE:
                    estados_a_saltear.append("No disponible")
                candidatas = candidatas[~candidatas[col_estado_prev].isin(estados_a_saltear)]

            print(f"\n🔎 Verificando {len(candidatas)} URLs guardadas que no aparecieron "
                  f"en la búsqueda de hoy (confirmando si siguen vivas o no)...")

            page_verif = await context.new_page()
            reconfirmadas = 0
            confirmadas_muertas = 0

            for _, fila_prev in candidatas.iterrows():
                url_prev = fila_prev[col_url_prev]
                if not isinstance(url_prev, str) or not url_prev.startswith("http"):
                    continue

                localidad_prev = fila_prev[col_localidad_prev] if col_localidad_prev and pd.notna(fila_prev.get(col_localidad_prev)) else ""

                exito = await navegar_con_reintento(page_verif, url_prev, reintentos=2, timeout=20000)
                if not exito:
                    confirmadas_muertas += 1
                    urls_confirmadas_muertas.add(url_prev)
                    print(f"  ❌ No disponible (no cargó la página): {url_prev}")
                    continue

                html_verif = await obtener_html_seguro(page_verif)
                soup_verif = BeautifulSoup(html_verif, "html.parser")

                texto_pagina = quitar_acentos(soup_verif.get_text(" "))
                señales_removida = [
                    "no se encuentra disponible",
                    "ya no esta disponible",
                    "fue eliminada",
                    "publicacion vencida",
                    "no encontrada",
                    "pagina no existe",
                ]
                tiene_titulo_valido = bool(soup_verif.select_one(".main-title h1, h1"))

                parece_removida = (
                    any(s in texto_pagina for s in señales_removida)
                    or not tiene_titulo_valido
                )

                if parece_removida:
                    confirmadas_muertas += 1
                    urls_confirmadas_muertas.add(url_prev)
                    print(f"  ❌ No disponible (confirmado en el sitio): {url_prev}")
                else:
                    datos_frescos = extraer_ficha_estricta(
                        soup_verif, url_prev, fecha_copia_actual, localidad_prev
                    )
                    lista_publicaciones.append(datos_frescos)
                    urls_procesadas.add(url_prev)
                    reconfirmadas += 1
                    print(f"  ✅ Sigue disponible: {datos_frescos['Título'][:40]}...")

                await page_verif.wait_for_timeout(300)

            await page_verif.close()
            print(f"\n📋 Verificación terminada: {reconfirmadas} reconfirmadas como vivas, "
                  f"{confirmadas_muertas} confirmadas como no disponibles.")

        await browser.close()

    # --- 3. CONSOLIDACIÓN: separamos solo las propiedades NUEVAS ---
    # Ya no armamos una única hoja gigante con todo mezclado. En vez de
    # eso: las que ya estaban en el Excel se dejan como estaban (excepto
    # las que confirmamos "No disponible" en el paso 2.5, que se
    # actualizan in-place más abajo). Acá solo nos quedamos con lo que
    # es genuinamente NUEVO para armar una hoja aparte.
    df_actual = pd.DataFrame(lista_publicaciones)

    if not df_previo.empty:
        col_url = "URL" if "URL" in df_previo.columns else df_previo.columns[0]
        urls_ya_conocidas = set(df_previo[col_url].dropna())
    else:
        urls_ya_conocidas = set()

    if not df_actual.empty:
        df_nuevas = df_actual[~df_actual["URL"].isin(urls_ya_conocidas)].copy()
    else:
        df_nuevas = df_actual

    if not df_nuevas.empty:
        df_nuevas["Estado"] = "Nueva"
        df_nuevas["Fecha de Copia"] = fecha_copia_actual
        df_nuevas["ID"] = [f"PROP-{i+1:03d}" for i in range(len(df_nuevas))]

    cols_orden = [
        "ID",
        "Estado",
        "Fecha de Copia",
        "Título",
        "Tipo de Propiedad",
        "Precio",
        "Localidad",
        "Dirección",
        "Descripción",
        "Dormitorios / Ambientes",
        "Baños",
        "Superficie",
        "Contacto / Vendedor",
        "Teléfono / WhatsApp",
        "Email",
        "URL"
    ]

    df_nuevas_final = df_nuevas.reindex(columns=cols_orden).copy() if not df_nuevas.empty else pd.DataFrame(columns=cols_orden)

    # --- 4. GUARDAR: actualizamos la hoja original IN-PLACE (solo el
    #     Estado de las que confirmamos caídas) y agregamos una hoja
    #     NUEVA (con fecha) solo con las propiedades nuevas de esta
    #     corrida. El resto del archivo y sus otras hojas quedan intactos.
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)

    fill_nueva = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_no_disp = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

    anchos_columnas = {
        "ID": 14,
        "Estado": 16,
        "Fecha de Copia": 22,
        "Título": 45,
        "Tipo de Propiedad": 24,
        "Precio": 18,
        "Localidad": 18,
        "Dirección": 32,
        "Descripción": 60,
        "Dormitorios / Ambientes": 24,
        "Baños": 14,
        "Superficie": 18,
        "Contacto / Vendedor": 30,
        "Teléfono / WhatsApp": 22,
        "Email": 30,
        "URL": 45
    }

    ALTO_FILA_FIJO = 28
    MAX_CHARS_DESC = 150

    def escribir_hoja(ws, df, fill_por_defecto=None):
        """Escribe encabezado + filas de un DataFrame en una hoja nueva,
        con el mismo formato que veníamos usando."""
        headers = list(df.columns)
        ws.append(headers)
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

        for i, row in df.iterrows():
            row_num = i + 2
            row_dict = row.to_dict()

            desc = str(row_dict.get("Descripción", ""))
            if pd.notna(desc) and len(desc) > MAX_CHARS_DESC:
                row_dict["Descripción"] = desc[:MAX_CHARS_DESC].strip() + "..."

            valores = [str(val) if pd.notna(val) else "" for val in row_dict.values()]
            ws.append(valores)
            ws.row_dimensions[row_num].height = ALTO_FILA_FIJO

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = data_font
                if fill_por_defecto:
                    cell.fill = fill_por_defecto
                cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=False)

        for col_idx, header in enumerate(headers, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = anchos_columnas.get(header, 22)

    if os.path.exists(ARCHIVO_EXCEL):
        wb = openpyxl.load_workbook(ARCHIVO_EXCEL)
    else:
        wb = openpyxl.Workbook()
        # Sacamos la hoja "Sheet" por defecto que crea openpyxl; la vamos
        # a reemplazar por la hoja de datos de verdad más abajo.
        wb.remove(wb.active)

    # 4a. Actualizamos in-place el Estado de las URLs confirmadas muertas,
    #     en TODAS las hojas de datos existentes (por si el usuario tiene
    #     más de una hoja con propiedades cargadas).
    if urls_confirmadas_muertas:
        hojas_actualizadas = 0
        for nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
            if ws.max_row < 2:
                continue
            encabezados = [c.value for c in ws[1]]
            if "URL" not in encabezados:
                continue
            col_url_idx = encabezados.index("URL") + 1
            col_estado_idx = encabezados.index("Estado") + 1 if "Estado" in encabezados else None
            if col_estado_idx is None:
                continue

            for row_num in range(2, ws.max_row + 1):
                valor_url = ws.cell(row=row_num, column=col_url_idx).value
                if valor_url in urls_confirmadas_muertas:
                    celda_estado = ws.cell(row=row_num, column=col_estado_idx)
                    celda_estado.value = "No disponible"
                    celda_estado.fill = fill_no_disp
                    hojas_actualizadas += 1
        print(f"\n📝 Actualizadas {hojas_actualizadas} filas a 'No disponible' "
              f"directamente en las hojas existentes.")

    # 4b. Si el archivo es nuevo (primera corrida), no hay hoja original
    #     todavía: creamos "Propiedades" con TODO lo scrapeado (que en ese
    #     caso es, por definición, todo nuevo).
    if "Propiedades" not in wb.sheetnames and df_previo.empty:
        ws_base = wb.create_sheet("Propiedades", 0)
        escribir_hoja(ws_base, df_nuevas_final, fill_por_defecto=fill_nueva)
        print(f"📁 Primera corrida: creada la hoja 'Propiedades' con {len(df_nuevas_final)} propiedades.")
    elif not df_nuevas_final.empty:
        # 4c. Corridas siguientes: las propiedades nuevas van a una hoja
        #     aparte con fecha, para no mezclar con la hoja original.
        nombre_hoja_nueva = "Nuevas_" + datetime.now().strftime("%Y-%m-%d_%H%M")
        nombre_hoja_nueva = nombre_hoja_nueva[:31]  # Excel limita a 31 caracteres
        sufijo = 1
        nombre_base = nombre_hoja_nueva
        while nombre_hoja_nueva in wb.sheetnames:
            sufijo += 1
            nombre_hoja_nueva = f"{nombre_base[:28]}_{sufijo}"
        ws_nueva = wb.create_sheet(nombre_hoja_nueva)
        escribir_hoja(ws_nueva, df_nuevas_final, fill_por_defecto=fill_nueva)
        print(f"\n🆕 Creada la hoja '{nombre_hoja_nueva}' con {len(df_nuevas_final)} propiedades nuevas.")
    else:
        print(f"\n🆕 No hubo propiedades nuevas en esta corrida: no se creó hoja nueva.")

    wb.save(ARCHIVO_EXCEL)
    print(f"\n✅ Proceso completado correctamente.")
    print(f"📊 Propiedades nuevas esta corrida: {len(df_nuevas_final)}")
    print(f"📊 URLs confirmadas como no disponibles esta corrida: {len(urls_confirmadas_muertas)}")
    print(f"📁 Guardado en '{os.path.abspath(ARCHIVO_EXCEL)}'")


if __name__ == "__main__":
    asyncio.run(run())